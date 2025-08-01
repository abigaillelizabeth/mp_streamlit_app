# IMPORTS STATEMENTS
import streamlit as st
import pandas as pd
import numpy as np
import csv
import re
import io 
from io import *
import os
from openpyxl import *
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
from openpyxl.styles import Font
import tempfile
from datetime import *
import toml
import zipfile
import openpyxl
import datetime

# ARENA METHODS  
# Function to reformat the input data
def process_mailing_data(input_file):
    # Read in the arena data
    raw_arena = pd.read_excel(input_file, sheet_name=0, header=None)
    # Drop any row that exactly matches the column headers (to remove repeated header rows)
    raw_arena = raw_arena[raw_arena.iloc[:, 0] != "Family Id"]

    # remove extra columns
    raw_arena = raw_arena.iloc[:, list(range(0, 15)) + [16]]
    #print(raw_arena.head())
    #print(raw_arena.head(3))
    # Group by Family Id (column 0) and Person ID (column 1)
    raw_arena.iloc[1:, 15] = raw_arena.iloc[1:, 15].astype(float)
    grouped_arena = raw_arena.groupby([0, 1], as_index=False).agg({
        2: 'first',    # 'Last Name'
        3: 'first',    # 'First Name'
        4: 'first',    # 'Nick Name'
        5: 'first',    # 'Donor Title'
        6: 'first',    # 'Spouse Title'
        7: 'first',    # 'Spouse Last Name' 
        8: 'first',    # 'Spouse First Name' 
        9: 'first',    # 'Spouse Nick Name'
        10: 'first',   # 'Address' 
        11: 'first',   # 'City'
        12: 'first',   # 'State' 
        13: 'first',   # 'Zip' 
        14: 'first',   # 'Email' 
        16: 'sum'      # 'Contribution Fund Amount' (sum)
    })

    # Rename the columns for clarity
    grouped_arena.columns = ['Family Id', 'Person ID', 'Last Name', 'First Name', 'Nick Name', 'Title', 
                             'Spouse Title', 'Spouse Last Name', 'Spouse First Name', 'Spouse Nick Name', 
                             'Address', 'City', 'State', 'Zip', 'Email', 'Total Contribution Fund Amount']
    # print(grouped_arena.shape)
    #print(grouped_arena.head(10))

    # Sort by Last name
    sorted_arena = grouped_arena.sort_values(by=['Last Name'])

    # Filter to persons & businesses
    blank_names = sorted_arena[(sorted_arena['First Name'].isna()) & (sorted_arena['Nick Name'].isna())]
    with_names = sorted_arena[~((sorted_arena['First Name'].isna()) & (sorted_arena['Nick Name'].isna()))]

     # Create header and blank row
    header_row = pd.DataFrame([list(sorted_arena.columns)], columns=sorted_arena.columns)
    blank_row = pd.DataFrame([[""] * len(sorted_arena.columns)], columns=sorted_arena.columns)

    final_df = pd.concat([
        #header_row,
        with_names,
        blank_row,
        header_row,
        blank_names,
    ], ignore_index=True)

    return final_df
# Function to generate the output data
def create_mailing_file(processed_arena_data, is_streamlit = True):
    if is_streamlit:
        # If running in Streamlit, keep the file in memory (no save to disk)
        arena_file = io.BytesIO()  # In-memory file for Streamlit (binary mode)
        processed_arena_data.to_excel(arena_file, index=False, engine='openpyxl')  # Use openpyxl for Excel output
        arena_file.seek(0)  # Go to the beginning of the in-memory file
    else:
        # If NOT running in Streamlit, save the file to disk
        output_file_path = "Sorted_Mailing_List.xlsx"
        processed_arena_data.to_excel(output_file_path, index=False, engine='openpyxl')  # Save to disk
        #processed_arena_data.to_excel(output_file_path, index=False)
        print(f"File successfully saved to {output_file_path}")
        arena_file = None  # Return None or any placeholder since file is saved on disk

    return arena_file
# Arena Main
def mainMailing(uploaded_file):
    processed_data = process_mailing_data(uploaded_file)
    print("data has been processed.")

    # Test the file creation function
    arena_final = process_mailing_data(processed_data, is_streamlit = False)
    #print(arena_final)

    if arena_final != None:
        print("File created successfully. Ready for download.")
    else:
        print("File has been saved to disk.")

    return arena_final
# Function to run arena methods  
def runMailing():
    st.header("Mailing List Report Upload")
    # Input Information
    uploaded_file = st.file_uploader("Upload an Arena-Downloaded Excel file", type="xlsx", key="arena_mailing")

    # Run the script when the button is pressed
    if st.button("Create Donor Summary"):
        if uploaded_file is not None:
            # Process the payroll data
            processed_data = process_mailing_data(uploaded_file)
            
            # Create the output file
            output_file = create_mailing_file(processed_data)

            st.success("Donor mailing list processed and ready for download!")

            # Provide download button for the payroll output
            st.download_button(
                label="Download Donor Summary",
                data=output_file,
                file_name="Sorted_Mailing_List.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                #mime="text/csv"
            )
        else:
            st.error("Please upload an arena file.")

# FIRST-TIME GIVERS METHODS
def process_ftg_data(input_file):
    # Read in the FTG data
    raw_ftg = pd.read_csv(input_file)
    #print(raw_ftg.head())

    return raw_ftg

def create_ftg_file(processed_ftg_data, is_streamlit=True):
    SHEET_1_NAME = "FTG Report"
    SHEET_2_NAME = "FTG Mailing Condensed"

    def format_ftg_report_sheet(ws):
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                # if isinstance(cell.value, (int, float)):
                #     cell.number_format = '#,##0.00'
                if isinstance(cell.value, str) and "date" in cell.value.lower():
                    cell.number_format = 'mm/dd/yyyy'


    if is_streamlit:
        # STEP 1: Write FTG Report to intermediate buffer
        intermediate = io.BytesIO()
        with pd.ExcelWriter(intermediate, engine='openpyxl') as writer:
            processed_ftg_data.to_excel(writer, index=False, sheet_name=SHEET_1_NAME)

        # ⚠️ Must reset before loading
        intermediate.seek(0)
        wb = load_workbook(intermediate)

        # STEP 2: Make edits
        if wb.sheetnames[0] != SHEET_1_NAME:
            wb.active.title = SHEET_1_NAME

        wb.create_sheet(SHEET_2_NAME)
        format_ftg_condensed_sheet(wb, SHEET_2_NAME)
        format_ftg_report_sheet(wb[SHEET_1_NAME])

        # STEP 3: Save final to NEW BytesIO buffer
        final_output = io.BytesIO()
        wb.save(final_output)

        # Must reset before returning
        final_output.seek(0)
        return final_output


    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                processed_ftg_data.to_excel(writer, index=False, sheet_name=SHEET_1_NAME)

            wb = load_workbook(tmp.name)

            if wb.sheetnames[0] != SHEET_1_NAME:
                wb.active.title = SHEET_1_NAME

            wb.create_sheet(SHEET_2_NAME)
            format_ftg_condensed_sheet(wb, SHEET_2_NAME)
            format_ftg_report_sheet(wb[SHEET_1_NAME])

            wb.save(tmp.name)
            return tmp.name

def format_ftg_condensed_sheet(wb, sheet_name="FTG Mailing Condensed"):
    source_sheet = wb["FTG Report"]
    target_sheet = wb[sheet_name]

    # Write new header
    headers = ["Person ID", "Full Name", "Address", "Contribution Amount"]
    target_sheet.append(headers)

    # Map headers to column indexes
    header_row = [str(cell.value).strip() for cell in next(source_sheet.iter_rows(min_row=1, max_row=1))]
    col_index = {col: i for i, col in enumerate(header_row)}

    # Group data by Person ID
    donor_data = defaultdict(lambda: {
        "full_name": "",
        "address": "",
        "amount": 0
    })

    for row in source_sheet.iter_rows(min_row=2, values_only=True):
        person_id = row[col_index["Person ID"]]
        first = row[col_index["First Name"]]
        last = row[col_index["Last Name"]]

        # Skip company entries
        if not first or not last:
            continue

        title = row[col_index["Donor Title"]] or ""
        full_name = f"{title} {first} {last}".strip()

        street = row[col_index["Address"]] or ""
        city = row[col_index["City"]] or ""
        state = row[col_index["State"]] or ""
        zip_code = row[col_index["Zip"]] or ""
        address = f"{street}, {city}, {state} {zip_code}".strip(", ")

        raw_amount = row[col_index["Contribution Amount"]]
        try:
            # Remove $ signs, commas, spaces
            cleaned = re.sub(r"[^\d.\-]", "", str(raw_amount))
            amount = float(cleaned)
        except (TypeError, ValueError):
            amount = 0


        donor_data[person_id]["full_name"] = full_name
        donor_data[person_id]["address"] = address
        donor_data[person_id]["amount"] += amount

    # Write one row per unique donor
    for person_id, info in donor_data.items():
        target_sheet.append([
            person_id,
            info["full_name"],
            info["address"],
            info["amount"]
        ])

    # Format amount column
    for row in target_sheet.iter_rows(min_row=2):
        amount_cell = row[3]
        if isinstance(amount_cell.value, (int, float)):
            amount_cell.number_format = '"$"#,##0.00'

    # Bold the header row
    for cell in target_sheet[1]:
        cell.font = Font(bold=True)

def mainFTG(uploaded_file):
    processed_data = process_ftg_data(uploaded_file)
    #print("data has been processed.")

    # Test the file creation function
    ftg_file = create_ftg_file(processed_data, is_streamlit = False)
    #print(arena_final)

    if ftg_file != None:
        print("File created successfully. Ready for download.")
    else:
        print("File has been saved to disk.")

    return ftg_file

def runFTG():
    st.header("First-Time Givers Report Upload")
    # Input Information
    uploaded_file = st.file_uploader("Upload a First-Time Givers report in CSV format", type=["csv"], key="ftg_file")

    # Run the script when the button is pressed
    if st.button("Create First-Time Givers Summary"):
        if uploaded_file is not None:
            # Process the payroll data
            processed_data = process_ftg_data(uploaded_file)
            
            # Create the output file
            output_file = create_ftg_file(processed_data)

            st.success("First-time givers list ready for download!")
            #st.write("Generated file size (bytes):", len(output_file.getvalue()))

            # Provide download button for the payroll output
            st.download_button(
                label="Download first-time givers summary",
                data=output_file,
                file_name="First_Time_Givers_dd.mm.yy.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Please upload a first-time givers file.")


# ASSURE CONVERSION METHODS
def process_assure_txt_input(input_txt):
    return FileExistsError

def create_assure_txt(input_txt):
    return FileExistsError

def process_assure_file(input_file):
    # Read in the Assure data
    raw_Assure = pd.read_excel(input_file, sheet_name=0)
    print(raw_Assure.head())
    print(raw_Assure.shape)

    # Drop rows that are completely empty
    raw_Assure.dropna(how='all', inplace=True)

    # Rename columns for consistency
    raw_Assure.columns = [col.strip() for col in raw_Assure.columns]

    # Keep only rows where either Debit or Credit has a value
    raw_Assure = raw_Assure[(raw_Assure["Debit"].notna()) | (raw_Assure["Credit"].notna())]

    # Fill NaNs with empty string or 0 where appropriate
    raw_Assure["Debit"] = raw_Assure["Debit"].fillna(0)
    raw_Assure["Credit"] = raw_Assure["Credit"].fillna(0)
    raw_Assure["Description"] = raw_Assure["Description"].fillna("")

    final_Assure = raw_Assure[~((raw_Assure["Debit"] > 0) & (raw_Assure["Credit"] > 0))]

    return final_Assure

def create_assure_txt(df, journal_date, accounting_period):
    output = io.StringIO()

    unused_field1 = "00000"
    co_num = "0001"
    fund_num = "00000"
    journal_type = "PR"
    journal_num = "00000"
    unused_field3 = "000"
    empty_field = ""
    date_str = journal_date
    lines = []
    for _, row in df.iterrows():
        # Parse amount (positive for debit, negative for credit)
        debit = row.get("Debit", 0)
        credit = row.get("Credit", 0)

        if pd.notna(debit) and float(debit) != 0:
            amount = int(round(float(debit) * 100))
        elif pd.notna(credit) and float(credit) != 0:
            amount = -int(round(float(credit) * 100))
        else:
            continue  # skip if neither debit nor credit present

        # Format field 2: full journal ID
        field2 = f"{co_num}{fund_num}{accounting_period.zfill(2)}{journal_type}{journal_num}"

        # Format Dept & Acct # (zero if not used)
        gl_number = str(row.get("GL Number", ""))
        if "-" in gl_number:
            dept, acct = gl_number.split("-")
            dept = dept.zfill(3)
            acct = acct.zfill(9)  # Pad account number to 6 digits (e.g. 5100 → 005100)
            account_full = f"{dept}{acct}"  # e.g. "021005100"
        else:
            account_full = "000000000"  # fallback if formatting is wrong

        # Build line of 9 fields
        line = [
            unused_field1,
            field2,
            unused_field3,
            date_str,
            #str(row.get("Description", "")),
            str(row.get("Description")),
            empty_field,
            account_full,
            str(amount),
            empty_field,
        ]

        lines.append(",".join(f'"{str(field)}"' for field in line))

    output.write("\r\n".join(lines))

        # ✅ Return as BytesIO (binary) or StringIO (text) depending on context
        # ✅ Return correct type depending on context
    if st:
        assure_file = io.BytesIO(output.getvalue().encode("utf-8"))
    else:
        assure_file = io.StringIO(output.getvalue())

    assure_file.seek(0)
    return assure_file

def mainAssure(uploaded_file):
    processed_assure = process_assure_file(uploaded_file)  # Use your cleaning logic
    print(processed_assure.head())
    print(processed_assure.shape)

    journal_date = "010125"
    accounting_period = "01"
    
    assure_final = create_assure_txt(processed_assure, journal_date, accounting_period)
    print(assure_final)
    print("File created successfully. Ready for download.")

    #output_path = "/Users/AbigailClark1/Documents/mp_work/data_scripts/python_scripts/Assure_JE_Output.txt"
    #with open(output_path, "wb") as f:
    #    f.write(assure_final.getvalue())
    #print(f"Assure JE file saved to: {output_path}")

    return assure_final

def runAssure():
    st.subheader("Assure GL File Upload")
    uploaded_file = st.file_uploader("Upload an excel assure file", type="xlsx", key="assure_upload")

    journal_date = st.text_input("Journal Date:", value="010125")
    accounting_period = st.text_input("Accounting Period:", value="01")

    # Run the script when the button is pressed
    if st.button("Generate Assure JE File"):
        if uploaded_file is not None:
            # Process the payroll data
            processed_data = process_assure_file(uploaded_file)
            
            # Create the output file
            output_file = create_assure_txt(processed_data, journal_date, accounting_period)
            st.success("Assure file processed and ready for download!")

            # Provide download button for the payroll output
            st.download_button(
                label="Download Assure Journal Entry",
                data=output_file,
                file_name="GLTRN2000.txt",
                mime="text/plain"
            )
        else:
            st.error("Please upload an Assure file.")


# PAYROLL METHODS
# Function to reformat the input data
def process_pr_data(input_file):
    # Read in the PR data
    raw_PR = pd.read_excel(input_file, sheet_name=1, header=None)
    #print(raw_PR.head())  # Print the first few rows to understand its structure
    #print(raw_PR.shape)   # Check the number of rows and columns

    # Find the index where "DEPT" appears in column 7 (index 6)
    dept_row_index = raw_PR[raw_PR[6] == "DEPT"].index[0]

    # Subset rows from the 'DEPT' row onward, and select columns 7 to 9 (indices 5-8)
    PR_1 = raw_PR.iloc[dept_row_index:, 6:10]
    #print(PR_1.head)  # Print the first few rows to understand its structure
    #print(PR_1.shape)   # Check the number of rows and columns

    # Set column names
    PR_1.columns = PR_1.iloc[0]
    PR_1.columns = PR_1.columns.str.strip()  # Remove whitespace
    PR_2 = PR_1  # Create PR_2

    # Drop rows with NaN values in 'Dept' and 'Acct'
    PR_3 = PR_2.dropna(subset=[PR_2.columns[0], PR_2.columns[1]], how='all')

    # Drop rows with NaN values in 'Debit' and 'Credit'
    PR_4 = PR_3.dropna(subset=[PR_3.columns[2], PR_3.columns[3]], how='all')

    # Step 5: Convert to numericals
    #PR_4['DEBITS'] = pd.to_numeric(PR_4['DEBITS'], errors='coerce')  # Convert DEBITS to numeric, coercing errors to NaN
    #PR_4['CREDITS'] = pd.to_numeric(PR_4['CREDITS'], errors='coerce')  # Convert CREDITS to numeric, coercing errors to NaN

    PR_4.loc[:, 'DEBITS'] = pd.to_numeric(PR_4['DEBITS'], errors='coerce')
    PR_4.loc[:, 'CREDITS'] = pd.to_numeric(PR_4['CREDITS'], errors='coerce')

    # Print out column data types (for debugging)
    #print(PR_4.dtypes)

    # Step 7: Check the file format
    required_columns = ['DEPT', 'ACCT', 'DEBITS', 'CREDITS']
    if not all(col in PR_4.columns for col in required_columns):
        raise ValueError("Input file must have 'Dept', 'Acct', 'Debit', and 'Credit' columns.")
    print("Passed required columns test. ")
    pr_data = PR_4

    return pr_data

# Function to generate the output data
def create_pr_txt (processed_pr_data, journal_date, accounting_period, description_1):
    print("now entering output file creation")
    # Creating debit lines
    debit_lines = processed_pr_data[processed_pr_data['DEBITS'].notna()].copy()  # Filter rows where DEBITS is not NaN
    debit_lines['Field1'] = "00000"
    debit_lines['Field2'] = "000100000" + accounting_period + "JE00000"
    debit_lines['Field3'] = "000"
    debit_lines['Field4'] = journal_date
    debit_lines['Field5'] = description_1
    debit_lines['Field6'] = ""
    debit_lines['Field7'] = np.where(debit_lines['DEPT'].astype(str) == "0",
                                  "00000000" + debit_lines['ACCT'].astype(str),
                                  "0" + debit_lines['DEPT'].astype(str) + "00000" + debit_lines['ACCT'].astype(str))
    debit_lines['Field8'] = debit_lines['DEBITS'].apply(lambda x: f"{x:.2f}")  # Format to 2 decimal places
    debit_lines['Field9'] = ""
    #print(debit_lines.dtypes)

    # Creating credit lines
    credit_lines = processed_pr_data[processed_pr_data['CREDITS'].notna()].copy()  # Filter rows where CREDITS is not NaN
    credit_lines['Field1'] = "00000"
    credit_lines['Field2'] = "000100000" + accounting_period + "JE00000"
    credit_lines['Field3'] = "000"
    credit_lines['Field4'] = journal_date
    credit_lines['Field5'] = description_1
    credit_lines['Field6'] = ""
    credit_lines['Field7'] = np.where(credit_lines['DEPT'].astype(str) == "0", 
                                  "00000000" + credit_lines['ACCT'].astype(str),
                                  "0" + credit_lines['DEPT'].astype(str) + "00000" + credit_lines['ACCT'].astype(str))
    credit_lines['Field8'] = credit_lines['CREDITS'].apply(lambda x: f"{x * -1:.2f}" if x != 0 else "0.00")
    credit_lines['Field9'] = ""
    #print(credit_lines.dtypes)

    # Binding all lines (combining debit and credit lines)
    gltrn_df = pd.concat([debit_lines, credit_lines], ignore_index=True)
    gltrn_df = gltrn_df.filter(regex="^Field")  # Select only the columns that start with "Field"

    # Save the output to a .txt file
    if st:
        pr_file = io.BytesIO()  # In-memory file for Streamlit (binary mode)
    else:
        pr_file = io.StringIO()  # In-memory file for terminal (text mode)
    gltrn_df.to_csv(pr_file, index=False, header=False, sep=",", quotechar='"', quoting=csv.QUOTE_ALL)
    pr_file.seek(0)  # Go to the beginning of the in-memory file

    return pr_file

# Payroll Main
def mainPR(uploaded_file):
    processed_data = process_pr_data(uploaded_file)
    print("data has been processed.")

    # Test the file creation function
    journal_date = "010125"
    accounting_period = "01"
    description_1 = "Payroll Entry"
    pr_final = create_pr_txt(processed_data, journal_date, accounting_period, description_1)

    # Print & save the output as needed
    print("File created successfully. Ready for download.")
    #print(pr_final)
    #output_content = pr_final.getvalue()
    #print(output_content)  # This will print the file content to the terminal

    return pr_final
# Function to run payroll methods  
def runPayroll():
    st.header("Payroll File Upload")
    # Input Information
    uploaded_file = st.file_uploader("Upload a completed Payroll Excel file", type="xlsx", key = "payroll_file")
    journal_date = st.text_input("Journal Date:", value="010125")
    accounting_period = st.text_input("Accounting Period:", value="01")
    description_1 = st.text_input("Description for Journal Entry:", value="Payroll Entry xx.xx.xx")
    
    # Run the script when the button is pressed
    if st.button("Generate Payroll JE File"):
        if uploaded_file is not None:
            # Process the payroll data
            processed_data = process_pr_data(uploaded_file)
            
            # Create the output file
            output_file = create_pr_txt(processed_data, journal_date, accounting_period, description_1)

            st.success("Payroll file processed and ready for download!")

            # Provide download button for the payroll output
            st.download_button(
                label="Download Payroll File",
                data=output_file,
                file_name="GLTRN2000.txt",
                mime="text/csv"
            )
        else:
            st.error("Please upload a payroll file.")


# CIGNA METHODS 
# Function to reformat the input data
def process_cig_data(input_file):
    # Read in the Cigna data
    raw_Cigna = pd.read_excel(input_file, sheet_name=3, header=None)
    #print("Raw Cigna Data:")
    #print(raw_Cigna.head())
    #print(raw_Cigna.shape)

    # Set start and end rows (based on the "Employee ID" text in the first column)
    start_row = raw_Cigna[raw_Cigna[0] == "Employee ID"].index[0]
    end_row = raw_Cigna[raw_Cigna[0].isna() & (raw_Cigna.index > start_row)].index[0] - 1

    # Crop the data between start and end rows
    cropped_cig = raw_Cigna.iloc[start_row:end_row + 1, :]
    #print("Cropped Cigna Data:")
    #print(cropped_cig.head())
    #print(cropped_cig.shape)

    # Set column names (first row in cropped data)
    cropped_cig.columns = cropped_cig.iloc[0]
    cropped_cig = cropped_cig.iloc[1:].reset_index(drop=True)  # Remove the first row
    #print("Named Cropped Data:")
    #print(cropped_cig.head())

    data_csv = cropped_cig

    # Delete unnecessary columns (adjust column names based on the actual data)
    data_csv = data_csv.drop(data_csv.columns[[2, 12]], axis=1)

    #print("Modified Data:")
    #print(data_csv.head())

    # Normalize Employee Name columns (to avoid issues with spaces/case differences)
    data_csv['Employee Name'] = data_csv['Employee Name'].str.strip().str.lower()
    #print(data_csv.columns)
    
    # Load the employee departments list and normalize Employee.Name column
    master_depts = pd.read_excel("employee_depts_master.xlsx")
    master_depts['Employee.Name'] = master_depts['Employee.Name'].str.strip().str.lower()
    master_depts.rename(columns={'Employee.Name': 'Employee Name'}, inplace=True)
    #print(master_depts.columns)

    # Merge the employee departments with the data (based on normalized Employee.Name)
    data_csv = data_csv.merge(master_depts, on="Employee Name", how="left")

    # Drop 'Subgrp ID' column and replace it with 'Dept.Acct'
    if 'Subgrp.ID' in data_csv.columns:
        data_csv.drop(columns=['Subgrp.ID'], inplace=True)
    # Now Dept.Acct is in the data_csv, it replaces Subgrp ID

    # Reorder columns (add 'Dept.Acct' at the correct place and make sure it's clean)
    data_csv = data_csv[['Employee Name', 'Dept.Acct'] + [col for col in data_csv.columns if col not in ['Employee Name', 'Dept.Acct']]]

    #print("Departmental Data:")
    #print(data_csv.head())

    cig_data = data_csv
    return cig_data  
# Function to generate the output data
def create_cig_txt(processed_cig_data, journal_date, accounting_period, description_1, credit_acct):
    # Summarizing the data by Dept.Acct
    summary_data = processed_cig_data.groupby('Dept.Acct').agg(
        Sum_Medical=('Medical', 'sum'),
        Premium=('Amount Due (1)', 'sum'),
        Claims_Allocated=('Claims Funding (3)', 'sum'),
        Total=('Total (4)', 'sum')
    ).reset_index()
    
    # Step 1: Summing the necessary columns
    totals_row = summary_data[['Sum_Medical', 'Premium', 'Claims_Allocated', 'Total']].sum()

    # Step 2: Manually set 'Dept.Acct' to 'TOTAL'
    totals_row['Dept.Acct'] = 'TOTAL'

    # Step 3: Convert the totals_row (which is a Series) into a DataFrame (transpose)
    totals_row_df = totals_row.to_frame().T

    # Step 4: Append the totals row to the summary_data DataFrame
    summary_data = pd.concat([summary_data, totals_row_df], ignore_index=True)
    #print("Summary Data:")
    #print(summary_data)

    # Creating debit lines
    debit_lines = summary_data[summary_data['Dept.Acct'] != "TOTAL"].copy()
    debit_lines['Field1'] = "00000"
    debit_lines['Field2'] = "000100000" + accounting_period + "RE00000"
    debit_lines['Field3'] = "000"
    debit_lines['Field4'] = journal_date
    debit_lines['Field5'] = description_1
    debit_lines['Field6'] = ""
    debit_lines['Field7'] = debit_lines['Dept.Acct'].apply(lambda x: f"{x[:3]}00000{x[-4:]}")
    #debit_lines['Field8'] = debit_lines['Total'].apply(lambda x: f"{x*100:.0f}")  # Format total as integer
    debit_lines['Field8'] = debit_lines['Total'].apply(lambda x: f"{round(x, 2):.2f}")  # Round to 2 decimal places
    debit_lines['Field9'] = ""

    # Creating credit lines
    total_sum = summary_data[summary_data['Dept.Acct'] == "TOTAL"]['Total'].iloc[0]
    credit_line = pd.DataFrame({
        'Field1': ["00000"],
        'Field2': [f"000100000{accounting_period}RE00000"],
        'Field3': ["000"],
        'Field4': [journal_date],
        'Field5': [description_1],
        'Field6': [""],
        'Field7': [f"00000000{credit_acct}"],
        #'Field8': [f"{-total_sum * 100:.0f}"],  # Negative for credit
        'Field8': [f"{round(-total_sum, 2):.2f}"],  # Round to 2 decimal places and make it negative for credit
        'Field9': [""]
    })

    # Combine debit and credit lines
    gltrn_df = pd.concat([debit_lines, credit_line], ignore_index=True)
    gltrn_df = gltrn_df.filter(regex="^Field")  # Keep only the "Field" columns
    #print("GLTRN Data:")
    #print(gltrn_df)

    # Save the output to a .txt file
    if st:
        cig_file = io.BytesIO()  # In-memory file for Streamlit (binary mode)
    else:
        cig_file = io.StringIO()  # In-memory file for terminal (text mode)
    gltrn_df.to_csv(cig_file, index=False, header=False, sep=",", quotechar='"', quoting=csv.QUOTE_ALL)
    cig_file.seek(0)  # Go to the beginning of the in-memory file

    return cig_file
# Cigna Main
def mainCig(uploaded_file):
    processed_data = process_cig_data(uploaded_file)
    print("data has been processed.")

    # Test the file creation function
    journal_date = "010125"
    accounting_period = "01"
    description_1 = "Cigna Entry"
    credit_acct = "1130"
    cig_final = create_cig_txt(processed_data, journal_date, accounting_period, description_1, credit_acct)

    # Print & save the output as needed
    print("File created successfully. Ready for download.")

    return cig_final
# Function to run cigna methods  
def runCigna():
    st.header("Cigna File Upload")
    # Input Information
    uploaded_file = st.file_uploader("Upload a Cigna-downloaded Excel file", type="xlsx", key = "cigna_file")
    journal_date = st.text_input("Journal Date:", value="010125")
    accounting_period = st.text_input("Accounting Period:", value="01")
    description_1 = st.text_input("Description for Journal Entry:", value="Cigna Entry xx.xx.xx")
    credit_acct = "1130"
    
    # Run the script when the button is pressed
    if st.button("Generate Cigna JE File"):
        if uploaded_file is not None:
            # Process the Cigna data
            processed_data = process_cig_data(uploaded_file)
            
            # Create the output file
            output_file = create_cig_txt(processed_data, journal_date, accounting_period, description_1, credit_acct)

            st.success("Cigna file processed and ready for download!")

            # Provide download button for the Cigna output
            st.download_button(
                label="Download Cigna File",
                data=output_file,
                file_name="GLTRN2000.txt",
                mime="text/csv"
            ) 
        else:
            st.error("Please upload a Cigna file.")

# CONTRIBUTIONS- MASTER ADDRESS MERGER
def master_address_importer(uploaded_person_report):
    df = pd.read_excel(uploaded_person_report)

    # Create full address column
    df["Full Address"] = (
    df["Street Address"].fillna('').astype(str) + ", " +
    df["City"].fillna('').astype(str) + ", " +
    df["State"].fillna('').astype(str) + " " +
    df["Zip Code"].fillna('').astype(str)
    )

    return df

def master_donor_file(uploaded_donor_report, uploaded_person_report):
    # Generate address_df from person report
    address_df = master_address_importer(uploaded_person_report)

    donor_df = pd.read_excel(uploaded_donor_report)

    # Merge using Person ID from address_df and person_id from donor_df
    merged = donor_df.merge(
        address_df,
        left_on="person_id",
        right_on="Person ID",
        how="left"
    )

    return merged

def runMasterAddressMerger():
    st.header("Donor File & Master Address Report Merger")

    # Simple, one-shot upload
    uploaded_person_report = st.file_uploader("Upload Master Person Export (obtained from Membership -> Public Lists -> Master Person Report)", 
        type="xlsx", key="address_upload")
    uploaded_donor_report = st.file_uploader("Upload Arena Donor Export (obtained from Contributions -> Contribution List -> Contributors)", 
        type="xlsx", key="donor_upload")

    # One button to do everything
    if st.button("Create Donor Report"):
        if uploaded_person_report is not None and uploaded_donor_report is not None:
            merged_df = master_donor_file(uploaded_donor_report, uploaded_person_report)

            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                merged_df.to_excel(writer, index=False, sheet_name="Merged Donor Report")
            output.seek(0)

            st.success("Report ready for download!")
            st.download_button(
                label="Download",
                data=output,
                file_name="Donor_Report_With_Addresses.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.error("Please upload both files.")





# CONTRIBUTIONS - ARENA

def arena_all_new(uploaded_arena_files):
    arena_data_list = []  # List to store the processed data from each file

    for idx, uploaded_arena_file in enumerate(uploaded_arena_files):
        filename = uploaded_arena_file.name.lower()
        # Read in the Arena data based on type
        if filename.endswith('.csv'):
            raw_contrib_arena = pd.read_csv(uploaded_arena_file)
        elif filename.endswith('.xlsx'):
            raw_contrib_arena = pd.read_excel(uploaded_arena_file, sheet_name=0, header=None)
        else:
            st.warning(f"Unsupported file type: {uploaded_arena_file.name}")
            continue

        #raw_contrib_arena = pd.read_excel(uploaded_arena_file, sheet_name=0, header=None)

        # Call the arena_col_names function to adjust the column headers
        raw_contrib_arena = arena_col_names(raw_contrib_arena)

        # Add a column for "Batch #" based on the file name (5-digit string)
        batch_number = uploaded_arena_file.name.split('.')[0]  # Get the 5-digit file name without extension
        raw_contrib_arena["Batch #"] = batch_number  # Add the "Batch #" column to the dataframe

        # Append the processed data to the list
        arena_data_list.append(raw_contrib_arena)
        
        # If it's the first file, initialize the combined dataframe
        if idx == 0:
            combined_arena_data = raw_contrib_arena
        else:
            # If it's not the first file, append it to the combined dataframe
            combined_arena_data = pd.concat([combined_arena_data, raw_contrib_arena], ignore_index=True)
        
    return combined_arena_data

def arena_col_names(raw_contrib_arena):
    # Manually combine the first two rows into one header row, but only if both rows have values
    new_columns = [
        f"{col1} {col2}" if pd.notna(col1) and pd.notna(col2) else col1 if pd.notna(col1) else col2
        for col1, col2 in zip(raw_contrib_arena.iloc[0], raw_contrib_arena.iloc[1])
    ]
    # Assign the new concatenated columns as the header
    raw_contrib_arena.columns = new_columns
    raw_contrib_arena = raw_contrib_arena.drop([0, 1])  # Drop the first two rows (header rows)
    # Move the data from row 1 (index 1) to the top (index 0)
    raw_contrib_arena.reset_index(drop=True, inplace=True)

    return raw_contrib_arena

def arena_merge(uploaded_arena_files):
    #print("running arena_merge")
    # Check if all files have a 5-digit name (without the file extension)
    all_five_digits = True
    
    # Iterate through each file in the uploaded files
    for arena_file in uploaded_arena_files:

        print(f"filename: {arena_file.name}")
        print(f"type: {type(arena_file)}")
        arena_file.seek(0)
        print(f"first 10 bytes: {arena_file.read(10)}")
        arena_file.seek(0)  # reset after peeking

        file_name = arena_file.name.split('.')[0]  # Remove file extension to check the name length
        if len(file_name) != 5 or not file_name.isdigit():  # Check if the length is not 5 or it's not numeric
            all_five_digits = False
            break
    
    # Call appropriate function based on the file name length
    # if all_five_digits:
    #     combined_arena_data = arena_all_new(uploaded_arena_files)
    # else:
    #     combined_arena_data = arena_master_included(uploaded_arena_files)
    combined_arena_data = arena_all_new(uploaded_arena_files)

    return combined_arena_data

def arena_excel(combined_arena_data):
    combined_arena_data["Contribution Date"] = pd.to_datetime(combined_arena_data["Contribution Date"], errors="coerce")
    # print("testing arena formatting")

    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Arena Details"

    # Write column headers
    ws.append(combined_arena_data.columns.tolist())

    # Write data rows
    for row in combined_arena_data.itertuples(index=False):
        ws.append(list(row))

    # Apply formatting
    headers = combined_arena_data.columns.tolist()
    for row in ws.iter_rows(min_row=2):
        for i, header in enumerate(headers):
            cell = row[i]
            header_lower = str(header).lower().strip()
            if header_lower in ["amount", "contribution amount", "total"]:
                cell.number_format = '"$"#,##0.00'
            elif "date" in header_lower:
                if isinstance(cell.value, datetime.datetime):
                    cell.number_format = 'MM/DD/YYYY'
                else:
                    cell.number_format = 'mm/dd/yy'
            else:
                cell.number_format = 'General'

    # Save workbook to a BytesIO stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def runArenaContributions():
    # Upload Arena Batch Files (Allow multiple files)
    uploaded_arena_files = st.file_uploader(
        "Upload Arena Batch Files ", type="xlsx", 
        accept_multiple_files=True, 
        key = "arena_batch_files"
        )
    # Process the uploaded Arena files
    if st.button("Import Arena Batches"):
        if uploaded_arena_files:
            # Call the process_contrib_arena to handle the file processing and combine the files
            combined_arena_data = arena_merge(uploaded_arena_files)

            # Store the combined Arena data in session state
            st.session_state.arena_data = combined_arena_data
            st.success("Arena batches processed and combined successfully")

            if combined_arena_data is not None and not combined_arena_data.empty:
                output = arena_excel(combined_arena_data)
                # st.download_button(
                #     label="Download Merged Arena Data",
                #     data=output,
                #     file_name="merged_arena_batches.xlsx",
                #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                # )
            else:
                st.error("No valid data to write to Excel.")
        else:
            st.error("Please upload at least one Arena batch.")
        return 

def arena_from_master(master_file):
    try:
        df = pd.read_excel(master_file, sheet_name="Arena Details")
        df = df.loc[:, df.columns.notna()]  # Remove any NaN column headers
        df = df[df.columns[df.columns != "Unnamed: 0"]]  # Remove row-number index if present
        return df
    except Exception as e:
        st.error(f"Failed to read 'Arena Details' from master file: {e}")
        return pd.DataFrame()

# CONTRIBUTIONS - EASY-TITHE
def ezt_merge(uploaded_ezt_data): 
    #print("running ezt_merge")
    ezt_data_list = []  # List to store the processed data from each file
    
    # Iterate through each file in the uploaded files
    for idx, ezt_file in enumerate(uploaded_ezt_data): 
        # Readin the EZT data based on type
        filename = ezt_file.name.lower()
        if filename.endswith('.csv'):
            raw_contrib_ezt = pd.read_csv(ezt_file)
        elif filename.endswith('.xlsx'):
            raw_contrib_ezt = pd.read_excel(ezt_file, sheet_name=0, header=0)
        else:
            continue  # Skip unsupported formats

        raw_contrib_ezt = raw_contrib_ezt[raw_contrib_ezt["Date"].notna()] #  Remove summary rows
        ezt_data_list.append(raw_contrib_ezt) # Append the processed data to the list

        if idx == 0: # If it's the first file, initialize the combined dataframe
            combined_ezt_data = raw_contrib_ezt 
        else: # If it's not the first file, append it to the combined dataframe
            combined_ezt_data = pd.concat([combined_ezt_data, raw_contrib_ezt], ignore_index=True)
    return combined_ezt_data

def ezt_excel(combined_ezt_data):
    wb = Workbook()
    ws = wb.active
    ws.title = "EZT Details"

    # Write headers
    ws.append(combined_ezt_data.columns.tolist())

    # Write data rows
    for row in combined_ezt_data.itertuples(index=False):
        ws.append(list(row))

    # Apply formatting
    headers = combined_ezt_data.columns.tolist()
    for row in ws.iter_rows(min_row=2):
        for i, header in enumerate(headers):
            cell = row[i]
            header_lower = str(header).lower().strip()
            if header_lower == "gross gift":
                cell.number_format = '"$"#,##0.00'
            elif header_lower == "date":
                cell.number_format = 'mm/dd/yy'
            else:
                cell.number_format = 'General'

    # Save workbook to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def runEZTContributions():
    #st.write("EasyTithe Batch File Upload")
    # Upload the EZT batch file
    uploaded_ezt_files = st.file_uploader(
        "Upload EasyTithe Batch Files ",
        type=["xlsx", "csv"],
        accept_multiple_files=True, 
        key="ezt_batch_files"
    )
    #Run the script when the button is clicked for EZT file
    if st.button("Import EasyTithe Batches"):
        if uploaded_ezt_files:
            # print("EZT Data Imported")
            # Call the ezt_contributions to handle the file processing and combine the files
            combined_ezt_data = ezt_merge(uploaded_ezt_files)
            #print("Combined EZT data")

            # Store the combined EZT data in session state
            st.session_state.ezt_data = combined_ezt_data

            st.success("EasyTithe batches processed and combined successfully")
        
            if combined_ezt_data is not None and not combined_ezt_data.empty:
                output = ezt_excel(combined_ezt_data)
                # st.download_button(
                #     label="Download Merged EZT Data",
                #     data=output,
                #     file_name="merged_EZT_batches.xlsx",
                #     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                # )
            else:
                st.error("No valid data to write to Excel.")
        else:
            st.error("Please upload at least one EasyTithe batch.")

def ezt_from_master(master_file):
    try:
        df = pd.read_excel(master_file, sheet_name="EZT Details")
        df = df.loc[:, df.columns.notna()]
        df = df[df.columns[df.columns != "Unnamed: 0"]]
        return df
    except Exception as e:
        st.error(f"Failed to read 'EZT Details' from master file: {e}")
        return pd.DataFrame()

# CONTRIBUTIONS - IMPORTING
def runAllImports(uploaded_arena_files, uploaded_ezt_files, uploaded_master_file):
    has_new_arena = False
    has_new_ezt = False

    # Handle Arena files
    if uploaded_arena_files:
        combined_arena_data = arena_merge(uploaded_arena_files)
        st.session_state.arena_data = combined_arena_data
        st.success("Arena batches processed successfully.")
        has_new_arena = True

    # Handle EZT files
    if uploaded_ezt_files:
        combined_ezt_data = ezt_merge(uploaded_ezt_files)
        st.session_state.ezt_data = combined_ezt_data
        st.success("EasyTithe batches processed successfully.")
        has_new_ezt = True

    # Handle master file (optional, only useful if combined with new batches)
    if uploaded_master_file:
        st.session_state.master_file = uploaded_master_file
        if not (has_new_arena or has_new_ezt):
            st.warning("Master file provided, but no new Arena or EZT batches were uploaded. No changes will be made.")
        else:
            st.success("Master workbook loaded successfully.")
    else:
        st.session_state.master_file = None

    # Give user clarity on what happened
    if not uploaded_arena_files and not uploaded_master_file:
        st.info("No Arena files uploaded.")

    if not uploaded_ezt_files and not uploaded_master_file:
        st.info("No EasyTithe files uploaded.")


# CONTRIBUTIONS - MATCHING BY ID
def reorder_merged_columns(merged, arena_df, ezt_df):
    arena_cols = list(arena_df.columns)
    ezt_cols = list(ezt_df.columns)

    arena_id_cols = ["Arena Transaction ID", "Arena Batch #"]
    ezt_id_cols = ["EZT Transaction ID", "EZT Batch ID"]

    arena_other_cols = [col for col in arena_cols if col not in ["Transaction Detail", "Batch #"]]
    ezt_other_cols = [col for col in ezt_cols if col not in ["Transaction Number", "Batch ID"]]

    arena_other_cols = [col for col in arena_other_cols if col not in arena_id_cols]
    ezt_other_cols = [col for col in ezt_other_cols if col not in ezt_id_cols]

    final_order = arena_id_cols + arena_other_cols + ezt_id_cols + ezt_other_cols
    return [col for col in final_order if col in merged.columns]

def matched_from_master(master_file):
    if master_file is None:
        return pd.DataFrame()

    try:
        wb = openpyxl.load_workbook(master_file, data_only=True)
        ws = wb["Matched Contributions"]
        data = ws.values

        # Skip first 8 rows (summary counts, title rows)
        for _ in range(8):
            next(data)

        # Extract header
        columns = next(data)
        rows = list(data)

        return pd.DataFrame(rows, columns=columns)
    except Exception as e:
        st.warning(f"Could not extract prior matched transactions: {e}")
        return pd.DataFrame()

def categorized_matches(match_by_id_df, unmatched_df, ezt_df):
    wb = Workbook()
    wb.remove(wb.active)

    # Summary counts
    match_count = len(match_by_id_df)
    arena_only_count = len(unmatched_df[unmatched_df["Match Type"] == "No Match Found: Arena only"])
    ezt_only_count = len(unmatched_df[unmatched_df["Match Type"] == "No Match Found: EZT only"])
    total_unmatched = arena_only_count + ezt_only_count

    # ------------------------
    # Sheet 1: Matched Contributions (Fully Matched EZT Batches only)
    # ------------------------
    ws_matched = wb.create_sheet(title="Matched Contributions")

    # Count full match status
    ezt_full_counts = ezt_df["EZT Batch ID"].astype(str).value_counts()
    matched_counts = match_by_id_df["EZT Batch ID"].astype(str).value_counts()

    # Determine fully matched batch IDs
    fully_matched_batches = [
        batch_id for batch_id in matched_counts.index
        if matched_counts[batch_id] == ezt_full_counts.get(batch_id, -1)
    ]
    # Write summary
    summary_rows = [
        ["Fully Matched EZT Batches", len(fully_matched_batches)],
        ["Total Matched Transactions", match_count],
        ["Arena Only", arena_only_count],
        ["EZT Only", ezt_only_count],
        ["Total Unmatched", total_unmatched],
        [],
    ]
    for row in summary_rows:
        ws_matched.append(row)

    # For each fully matched batch, write block
    for batch_id in fully_matched_batches:
        # Filter and sort
        subset = match_by_id_df[match_by_id_df["EZT Batch ID"].astype(str) == batch_id]
        subset = subset.sort_values(by="Arena Batch #")

        # Spacer row before each batch
        ws_matched.append([])

        # Clean up batch ID for header
        try:
            batch_id_clean = str(int(float(batch_id)))
        except:
            batch_id_clean = str(batch_id).strip()

        # Write bold section header
        section_row = ws_matched.max_row + 1
        ws_matched.append([f"EZT Batch #{batch_id_clean} - Fully Matched"])
        header_cell = ws_matched.cell(row=section_row, column=1)
        header_cell.font = Font(bold=True)

        # Write column headers (repeated per group)
        ws_matched.append(subset.columns.tolist())

        # Write data rows
        for row in subset.itertuples(index=False):
            ws_matched.append(list(row))

        # Blank line after group
        ws_matched.append([])

    # ------------------------
    # Sheet 2: Unmatched Transactions
    # ------------------------
    ws_unmatched = wb.create_sheet(title="Unmatched Transactions")

    # Sort if possible
    if "Arena Batch #" in unmatched_df.columns:
        unmatched_df = unmatched_df.sort_values(by="Arena Batch #", na_position="last")

    ws_unmatched.append(unmatched_df.columns.tolist())
    for row in unmatched_df.itertuples(index=False):
        ws_unmatched.append(list(row))

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# CONTRIBUTIONS - EXPORTING
def export_matched_excel(arena_df, ezt_df, prior_matched=pd.DataFrame()):
    # Make copies to avoid changing original data
    arena_df = arena_df.copy()
    ezt_df = ezt_df.copy()

    # Rename key columns for clarity
    arena_df = arena_df.rename(columns={
        "Transaction Detail": "Arena Transaction ID",
        "Batch #": "Arena Batch #"
    })
    ezt_df = ezt_df.rename(columns={
        "Transaction Number": "EZT Transaction ID",
        "Batch ID": "EZT Batch ID"
    })
    
    # Safe normalization: cast to int first to strip decimal, then to str
    arena_df["Arena Transaction ID"] = arena_df["Arena Transaction ID"].astype(float).astype(int).astype(str).str.strip()
    ezt_df["EZT Transaction ID"] = ezt_df["EZT Transaction ID"].astype(float).astype(int).astype(str).str.strip()

    merged = pd.merge(
        arena_df,
        ezt_df,
        left_on="Arena Transaction ID",
        right_on="EZT Transaction ID",
        how="outer",
        indicator=True
    )

    # Extract match categories
    match_by_transaction_id = merged[merged["_merge"] == "both"].copy()
    match_by_transaction_id["Arena Batch #"] = match_by_transaction_id["Arena Batch #"].astype(str)
    match_by_transaction_id = match_by_transaction_id.sort_values(by="Arena Batch #", ascending=True)
    arena_only = merged[merged["_merge"] == "left_only"].copy()
    ezt_only = merged[merged["_merge"] == "right_only"].copy()

    # Add Match Type Indicator Column
    match_by_transaction_id["Match Type"] = "Match Found"
    arena_only["Match Type"] = "No Match Found: Arena only"
    ezt_only["Match Type"] = "No Match Found: EZT only"

    # Drop the merge indicator before output
    for df in [match_by_transaction_id, arena_only, ezt_only]:
        df.drop(columns=["_merge"], inplace=True)

    # Apply consistent column order
    ordered_cols = reorder_merged_columns(merged, arena_df, ezt_df) + ["Match Type"]
    match_by_transaction_id = match_by_transaction_id[ordered_cols]
    arena_only = arena_only[ordered_cols]
    ezt_only = ezt_only[ordered_cols]

    # ⬅️ If a prior master was uploaded, preserve already-matched transactions
    if not prior_matched.empty:
        # Normalize types to ensure clean comparison
        prior_matched["Arena Transaction ID"] = prior_matched["Arena Transaction ID"].astype(str).str.strip()
        prior_matched["EZT Transaction ID"] = prior_matched["EZT Transaction ID"].astype(str).str.strip()
        match_by_transaction_id["Arena Transaction ID"] = match_by_transaction_id["Arena Transaction ID"].astype(str).str.strip()
        match_by_transaction_id["EZT Transaction ID"] = match_by_transaction_id["EZT Transaction ID"].astype(str).str.strip()

        # Build keys to find overlap
        prior_keys = prior_matched[["Arena Transaction ID", "EZT Transaction ID"]].dropna()
        current_keys = match_by_transaction_id[["Arena Transaction ID", "EZT Transaction ID"]]

        # Identify which matches are new (i.e., not already present in prior master)
        combined = match_by_transaction_id.merge(
            prior_keys,
            on=["Arena Transaction ID", "EZT Transaction ID"],
            how="left",
            indicator=True
        )
        new_matches = combined[combined["_merge"] == "left_only"].drop(columns=["_merge"])

        # Add back preserved matches from the master
        match_by_transaction_id = pd.concat([prior_matched, new_matches], ignore_index=True)

    # Build categorized Excel with labeled sections
    return categorized_matches(
        match_by_id_df=match_by_transaction_id,
        unmatched_df=pd.concat([arena_only, ezt_only], ignore_index=True),
        ezt_df = ezt_df  # <-- NEW
    )

def export_combined_excel(arena_df, ezt_df):

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    def write_and_format_sheet(df, title):
        ws = wb.create_sheet(title=title)
        ws.append(df.columns.tolist())
        for row in df.itertuples(index=False):
            ws.append(list(row))

        headers = df.columns.tolist()
        for row in ws.iter_rows(min_row=2):
            for i, header in enumerate(headers):
                cell = row[i]
                header_lower = str(header).lower().strip()
                if header_lower in ["gross gift", "amount", "contribution amount", "total"]:
                    cell.number_format = '"$"#,##0.00'
                elif "date" in header_lower:
                    cell.number_format = "mm/dd/yy"
                else:
                    cell.number_format = "General"

    # Add both sheets
    write_and_format_sheet(arena_df, "Arena Details")
    write_and_format_sheet(ezt_df, "EZT Details")

    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_contributions_master(arena_df, ezt_df):
    def apply_formatting(ws):
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2):
            for i, header in enumerate(headers):
                cell = row[i]
                header_lower = str(header).lower().strip()
                if header_lower in ["gross gift", "amount", "contribution amount", "total"]:
                    cell.number_format = '"$"#,##0.00'
                elif "date" in header_lower:
                    cell.number_format = "mm/dd/yy"
                else:
                    cell.number_format = "General"

    def copy_sheet_with_formatting(source_ws, target_wb, sheet_name):
        new_ws = target_wb.create_sheet(title=sheet_name)
        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
                new_cell.number_format = cell.number_format

    # Generate formatted Excel files in memory
    arena_output = arena_excel(arena_df)
    ezt_output = ezt_excel(ezt_df)
    #matched_output = export_matched_excel(arena_df, ezt_df)
    prior_matched = matched_from_master(st.session_state.get("master_file")) if "master_file" in st.session_state else pd.DataFrame()
    matched_output = export_matched_excel(arena_df, ezt_df, prior_matched)

    # Load formatted workbooks
    arena_output.seek(0)
    arena_wb = load_workbook(arena_output)
    ezt_output.seek(0)
    ezt_wb = load_workbook(ezt_output)
    matched_output.seek(0)
    matched_wb = load_workbook(matched_output)

    # try:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet in matched_wb.sheetnames:
        apply_formatting(matched_wb[sheet])
        copy_sheet_with_formatting(matched_wb[sheet], wb, sheet)

    for sheet in arena_wb.sheetnames:
        apply_formatting(arena_wb[sheet])
        copy_sheet_with_formatting(arena_wb[sheet], wb, sheet)

    for sheet in ezt_wb.sheetnames:
        apply_formatting(ezt_wb[sheet])
        copy_sheet_with_formatting(ezt_wb[sheet], wb, sheet)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def runMatchContributions():
    # arena_ready = 'arena_data' in st.session_state
    # ezt_ready = 'ezt_data' in st.session_state

    # if arena_ready and ezt_ready:
    #     st.header("Contribution Report Download Options")
    #     st.write("Click the button below to generate the full contribution report.")

    #     if st.button("Generate Reports"):
    master_file = st.session_state.get("master_file", None)
    arena_ready_from_master = False
    ezt_ready_from_master = False

    if master_file:
        try:
            prior_arena = arena_from_master(master_file)  # <- this is the function
            prior_ezt = ezt_from_master(master_file)
            arena_ready_from_master = not prior_arena.empty
            ezt_ready_from_master = not prior_ezt.empty
        except Exception as e:
            st.warning(f"Could not load sheets from master file: {e}")

    arena_ready = 'arena_data' in st.session_state or arena_ready_from_master
    ezt_ready = 'ezt_data' in st.session_state or ezt_ready_from_master


    # ✅ Enforce: must have both Arena and EZT (either from new upload or master)
    if arena_ready and ezt_ready:
        st.header("Contribution Report Download Options")
        st.write("Click the button below to generate the full contribution report.")

        if st.button("Generate Reports"):
            # Load prior master data if provided
            uploaded_master = st.session_state.get("master_file", None)

            if uploaded_master:
                prior_arena = arena_from_master(uploaded_master)
                prior_ezt = ezt_from_master(uploaded_master)
            else:
                prior_arena = pd.DataFrame()
                prior_ezt = pd.DataFrame()


            # Remove duplicates before combining
            #current_arena = st.session_state.arena_data
            current_arena = st.session_state.get("arena_data", pd.DataFrame())
            #current_ezt = st.session_state.ezt_data
            current_ezt = st.session_state.get("ezt_data", pd.DataFrame())
            
            # Normalize column names and transaction values
            prior_arena.columns = [col.strip() for col in prior_arena.columns]
            current_arena.columns = [col.strip() for col in current_arena.columns]

            if "Transaction Detail" in prior_arena.columns and "Transaction Detail" in current_arena.columns:
                prior_arena["Transaction Detail"] = prior_arena["Transaction Detail"].astype(str).str.strip()
                current_arena["Transaction Detail"] = current_arena["Transaction Detail"].astype(str).str.strip()

                # Drop duplicates from prior_arena so they can be replaced by new ones
                prior_arena = prior_arena[~prior_arena["Transaction Detail"].isin(current_arena["Transaction Detail"])]


            # Only keep EZT rows with Transaction Number not in prior_ezt
            if (
                not prior_ezt.empty and 
                "Transaction Number" in prior_ezt.columns and
                not current_ezt.empty and 
                "Transaction Number" in current_ezt.columns
            ):
                current_ezt = current_ezt[
                    ~current_ezt["Transaction Number"].isin(prior_ezt["Transaction Number"])
                ]

            # Combine prior and current data
            combined_arena = pd.concat([prior_arena, current_arena], ignore_index=True)
            # Sort Arena data by Arena Batch #
            if "Batch #" in combined_arena.columns:
                combined_arena["Batch #"] = combined_arena["Batch #"].astype(str).str.strip()
                combined_arena = combined_arena.sort_values(by="Batch #", ascending=True)

            combined_ezt = pd.concat([prior_ezt, current_ezt], ignore_index=True)

            # Generate master Excel file with all 3 sheets
            master_excel = export_contributions_master(combined_arena, combined_ezt)

            st.download_button(
                label="Download Master Workbook (.xlsx)",
                data=master_excel,
                file_name="master_contributions_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        # st.info("Upload Arena and EZT batches to generate the full report.")
        st.info("You must upload both Arena and EasyTithe files (or upload one and combine it with a master file) to generate the report.")

def runContributions():
    st.header("Contribution Reports Processing")
    # Upload Prior Master Workbook (Optional)
    uploaded_master_file = st.file_uploader(
        "Upload Prior Master Workbook (Optional)",
        type=["xlsx"],
        key="existing_master_upload"
    )
    # Upload Arena Batch Files
    uploaded_arena_files = st.file_uploader(
        "Upload Arena Batch Files", 
        type=["xlsx", "csv"], 
        accept_multiple_files=True, 
        key="arena_batch_files"
    )
    # Upload EZT Batch Files
    uploaded_ezt_files = st.file_uploader(
        "Upload EasyTithe Batch Files", 
        type=["xlsx", "csv"], 
        accept_multiple_files=True, 
        key="ezt_batch_files"
    )
    # Unified Import Button
    if st.button("Import All Files"):
        runAllImports(uploaded_arena_files, uploaded_ezt_files, uploaded_master_file)

    # Continue to matching and export
    runMatchContributions()



# STREAMLIT METHODS
# Authenticate user
def authenticate(username, password):
    # Loop through the credentials to match the username and password
    for user in st.secrets["credentials"]["user"]:
        if user["username"] == username and user["password"] == password:
            return True  # Successful authentication
    return False  # Failed authentication

def call_methods():
    st.title("MP File Import & Conversion App")

    with st.sidebar:
        if st.button("🔒 Logout"):
            st.session_state.logged_in = False
            st.session_state.selected_function = None
            st.rerun()
        
        st.markdown("---")
        st.header("Tools Menu")

        st.markdown("### 📂 Contributions")
        if st.button("🧾 Contribution Reports"):
            st.session_state.selected_function = "Contribution Reports"

        st.markdown("### 📋 Arena List Formatting")
        if st.button("📨 Arena Mailing Report"):
            st.session_state.selected_function = "Arena Mailing Report"
        if st.button("👤 First-Time Givers Report"):
            st.session_state.selected_function = "First-Time Givers Report"
        if st.button("📍 Donor Address Merger"):
            st.session_state.selected_function = "Master Test Report"

        st.markdown("### 🧮 Journey Entry Creation")
        if st.button("💼 Payroll Workbook"):
            st.session_state.selected_function = "Payroll Workbook"
        if st.button("💊 Cigna Workbook"):
            st.session_state.selected_function = "Cigna Workbook"
        if st.button("⏳Assure Download"):
            st.session_state.selected_function = "Assure Download"

    # Show the selected tool
    selected_function = st.session_state.get("selected_function")

    if selected_function == "Contribution Reports":
        runContributions()
    elif selected_function == "Master Test Report":
        runMasterAddressMerger()
    elif selected_function == "Arena Mailing Report":
        runMailing()
    elif selected_function == "First-Time Givers Report":
        runFTG()
    elif selected_function == "Payroll Workbook":
        runPayroll()
    elif selected_function == "Cigna Workbook":
        runCigna()
    elif selected_function == "Assure Download":
        runAssure()

# Set streamlit logic 
def run_gui():
    # Set up session state to track login status
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False

    # Show login form if not logged in
    if not st.session_state.logged_in:
        st.title("MP File Conversion Login")
        st.write("Please enter your username and password.")
        
        # Create input fields for username and password
        username = st.text_input("Username")
        password = st.text_input("Password", type="password")
        login_button = st.button("Login")

        if login_button:
            if authenticate(username, password):
                # Successful login
                st.session_state.logged_in = True  # Store login status
                st.success("Login successful!")
                st.rerun()  # This is to refresh the page and hide the login form
            else:
                # Invalid credentials
                st.error("Invalid credentials. Please try again.")
    
    else:
        call_methods()

# Streamit running
if __name__ == "__main__":
    run_gui() # WITH AUTH
    #call_methods() # WITHOUT AUTH
    

# TERMINAL TESTING
# if __name__ == "__main__":
#     # PAYROLL
#     uploaded_PR = 'PR Journal Entry_03.25.2025-1.xlsx'  # Replace with the path to your test file
#     mainPR(uploaded_PR)
#     # CIGNA
#     uploaded_Cig = 'GroupPremiumStatementRpt_03.2025.xlsx'  # Replace with the path to your test file
#     mainCig(uploaded_Cig)
#     # ARENA 
#     uploaded_arena = 'DonorMailingTester.xlsx'  # Replace with the path to your test file
#     mainMailing(uploaded_arena)
#     # ASSURE
#     uploaded_Assure = 'Assure Test 06.03.2025-1.xlsx' # Replace with the path to your test file
#     mainAssure(uploaded_Assure)
